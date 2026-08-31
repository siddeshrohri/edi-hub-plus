"""
EDI Hub+ Pipeline — Evaluation Script
======================================
Compares Stage 6 predicted tags against ground truth annotations from the Excel.

For each dimension and each tag in the taxonomy, computes:
  TP  — predicted AND in ground truth
  TN  — not predicted AND not in ground truth
  FP  — predicted but NOT in ground truth
  FN  — not predicted but IS in ground truth

Then derives per-dimension Precision, Recall and F1.
Also reports Exact Match rate (predicted set == ground truth set) per dimension.

Usage:
    python eval.py --tagged tagged_39.json tagged_100.json ...
    python eval.py --tagged-dir ./tagged_outputs/
    python eval.py --tagged-dir ./tagged_outputs/ --excel ground_truth.xlsm
    python eval.py --tagged-dir ./tagged_outputs/ --dims individual_characteristics career_pathway
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from collections import defaultdict

import openpyxl

# ---------------------------------------------------------------------------
# Taxonomy (must match stage6_tagger.py)
# ---------------------------------------------------------------------------

TAXONOMY = {
    "resource_type": [
        "Academic Paper", "Blog Post", "Book Chapter", "Case Study/Testimonial",
        "EDI Intervention", "Evidence Document", "Guidance", "Legislation",
        "People Profile", "Policy Document", "Presentation or Poster",
        "Report or Article", "Resource Database", "Training Material",
    ],
    "resource_format": [
        "Database", "Dataset", "Document", "Mixed format",
        "Podcast", "Video", "Webinar", "Webpage",
    ],
    "individual_characteristics": [
        "Age", "Career Stage", "Caring Responsibilities", "Disability",
        "Education Status", "Gender", "Intersectionality", "Job Roles",
        "Linguistic", "Marriage and Civil Partnership", "Nationality",
        "Neurodiversity", "Physical Characteristics", "Pregnancy and Maternity",
        "Race and Ethnicity", "Religion and Belief", "Sex", "Sexual Orientation",
        "Socio-Economic Disadvantage", "Temporary Impairment", "Trans Identity",
        "Not Specified",
    ],
    "career_pathway": [
        "Career Mobility", "Career Progression", "Early Career", "Leadership",
        "Mentoring", "Mid-career", "Networking and Collaboration", "Recruitment",
        "Retention", "Visibility and Recognition",
    ],
    "organisational_culture": [
        "Accessibility", "Allyship and Advocacy", "Belonging and Engagement",
        "Bias", "Cultural Competency", "Data Collection and Monitoring",
        "Harassment Bullying and Microaggression", "Inclusive Leadership",
        "Inclusivity", "Representation", "Research Culture", "Research Design",
        "Strategy or Policy", "Work-Life Balance",
    ],
    "research_funding_process": [
        "Advertising", "Application Process", "Assessment Process",
        "Bias Mitigation", "Data Collection and Monitoring", "Eligibility Criteria",
        "Funding Guidance", "Pre and Post Award Support", "University Selection",
    ],
    "adoption_readiness_level": [
        "Initial Design", "Needs Identification", "Not applicable",
        "Pilot in Single Setting", "Proof of Concept",
        "Scaling in Multiple Settings", "Widespread Adoption",
    ],
}

# Map Excel column names → dimension keys
EXCEL_COL_TO_DIM = {
    "Resource Type":              "resource_type",
    "Resource Format ":           "resource_format",
    "Individual Characteristics": "individual_characteristics",
    "Career Pathway":             "career_pathway",
    "Organisational Culture":     "organisational_culture",
    "Research Funding Process":   "research_funding_process",
    "Adoption Readiness Level":   "adoption_readiness_level",
}

# ---------------------------------------------------------------------------
# Ground truth loader
# ---------------------------------------------------------------------------

def load_ground_truth(excel_path: str) -> dict:
    """
    Returns {resource_id: {dimension: set_of_tags}}.
    Tags are normalised to lowercase stripped strings for comparison.
    """
    wb = openpyxl.load_workbook(excel_path, keep_vba=True, data_only=True)
    ws = wb["Resources-to-show"]

    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ground_truth = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        rid  = str(data.get("Resource ID", "")).strip()
        if not rid or rid == "None":
            continue

        dims = {}
        for col_name, dim_key in EXCEL_COL_TO_DIM.items():
            raw = data.get(col_name, "") or ""
            if isinstance(raw, str) and raw.strip():
                tags = {t.strip().lower() for t in raw.split(",") if t.strip()}
            else:
                tags = set()
            dims[dim_key] = tags

        ground_truth[rid] = dims

    return ground_truth


# ---------------------------------------------------------------------------
# Predicted tag loader
# ---------------------------------------------------------------------------

def load_predictions(tagged_files: list[str]) -> dict:
    """
    Returns {resource_id: {dimension: set_of_tags}}.
    Tags normalised to lowercase for comparison.
    Merges multiple tagged files for the same resource (stage3 + stage4 + stage5).
    """
    predictions = defaultdict(lambda: defaultdict(set))

    for path in tagged_files:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)

        rid = str(data.get("resource_id", "")).strip()
        if not rid:
            print(f"  [WARN] No resource_id in {path} — skipping")
            continue

        for dim, tag_list in data.get("tags", {}).items():
            for tag_entry in tag_list:
                tag = tag_entry.get("tag", "").strip().lower()
                if tag:
                    predictions[rid][dim].add(tag)

    return {rid: dict(dims) for rid, dims in predictions.items()}


# ---------------------------------------------------------------------------
# Per-resource, per-dimension evaluation
# ---------------------------------------------------------------------------

def evaluate_dimension(
    pred_tags: set,
    gt_tags:   set,
    all_tags:  list,
) -> dict:
    """
    Given predicted tags and ground truth tags for one dimension on one resource,
    compute TP/TN/FP/FN per tag, then aggregate.

    Returns:
        {
            "tp": int, "tn": int, "fp": int, "fn": int,
            "exact_match": bool,
            "precision": float, "recall": float, "f1": float,
            "per_tag": {tag: {"tp":..., "fp":..., "fn":..., "tn":...}}
        }
    """
    # Normalise everything to lowercase for comparison
    pred_lower = {t.lower() for t in pred_tags}
    gt_lower   = {t.lower() for t in gt_tags}
    all_lower  = [t.lower() for t in all_tags]

    per_tag = {}
    tp = tn = fp = fn = 0

    for tag in all_lower:
        predicted = tag in pred_lower
        actual    = tag in gt_lower

        if predicted and actual:
            per_tag[tag] = {"tp": 1, "fp": 0, "fn": 0, "tn": 0}
            tp += 1
        elif predicted and not actual:
            per_tag[tag] = {"tp": 0, "fp": 1, "fn": 0, "tn": 0}
            fp += 1
        elif not predicted and actual:
            per_tag[tag] = {"tp": 0, "fp": 0, "fn": 1, "tn": 0}
            fn += 1
        else:
            per_tag[tag] = {"tp": 0, "fp": 0, "fn": 0, "tn": 1}
            tn += 1

    precision    = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    recall       = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    f1           = (2 * precision * recall / (precision + recall)
                    if (precision + recall) > 0 else 0.0)
    exact_match  = pred_lower == gt_lower

    return {
        "tp": tp, "tn": tn, "fp": fp, "fn": fn,
        "exact_match": exact_match,
        "precision": round(precision, 4),
        "recall":    round(recall, 4),
        "f1":        round(f1, 4),
        "per_tag":   per_tag,
    }


# ---------------------------------------------------------------------------
# Full evaluation across all resources and dimensions
# ---------------------------------------------------------------------------

def run_evaluation(
    predictions:  dict,
    ground_truth: dict,
    dims_filter:  list = None,
) -> dict:
    """
    Runs evaluation for every resource that has both a prediction and a ground truth.
    Returns a results dict with per-resource, per-dimension scores and overall summary.
    """
    dims_to_eval = dims_filter or list(TAXONOMY.keys())
    results      = {}

    matched_ids = set(predictions.keys()) & set(ground_truth.keys())
    only_pred   = set(predictions.keys()) - set(ground_truth.keys())
    only_gt     = set(ground_truth.keys()) - set(predictions.keys())

    if only_pred:
        print(f"  [WARN] {len(only_pred)} resource(s) have predictions but no ground truth: {sorted(only_pred)}")
    if only_gt:
        print(f"  [INFO] {len(only_gt)} resource(s) have ground truth but no predictions yet: {sorted(only_gt)}")

    for rid in sorted(matched_ids, key=lambda x: int(x) if x.isdigit() else x):
        pred = predictions[rid]
        gt   = ground_truth[rid]
        results[rid] = {}

        for dim in dims_to_eval:
            if dim not in TAXONOMY:
                continue
            pred_tags = pred.get(dim, set())
            gt_tags   = gt.get(dim, set())
            results[rid][dim] = evaluate_dimension(pred_tags, gt_tags, TAXONOMY[dim])

    return {
        "per_resource":  results,
        "matched_ids":   sorted(matched_ids),
        "unmatched_pred": sorted(only_pred),
        "unmatched_gt":   sorted(only_gt),
    }


# ---------------------------------------------------------------------------
# Aggregate summary per dimension
# ---------------------------------------------------------------------------

def summarise(eval_results: dict, dims_filter: list = None) -> dict:
    """
    Aggregates TP/TN/FP/FN across all resources per dimension.
    Computes macro Precision, Recall, F1 and Exact Match rate.
    """
    dims = dims_filter or list(TAXONOMY.keys())
    summary = {}

    for dim in dims:
        agg = {"tp": 0, "tn": 0, "fp": 0, "fn": 0, "exact_matches": 0, "total": 0}

        for rid, dim_scores in eval_results["per_resource"].items():
            if dim not in dim_scores:
                continue
            s = dim_scores[dim]
            agg["tp"]   += s["tp"]
            agg["tn"]   += s["tn"]
            agg["fp"]   += s["fp"]
            agg["fn"]   += s["fn"]
            agg["total"] += 1
            if s["exact_match"]:
                agg["exact_matches"] += 1

        tp, fp, fn = agg["tp"], agg["fp"], agg["fn"]
        precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
        recall    = tp / (tp + fn) if (tp + fn) > 0 else 0.0
        f1        = (2 * precision * recall / (precision + recall)
                     if (precision + recall) > 0 else 0.0)
        exact_rate = agg["exact_matches"] / agg["total"] if agg["total"] > 0 else 0.0

        summary[dim] = {
            **agg,
            "precision":   round(precision, 4),
            "recall":      round(recall, 4),
            "f1":          round(f1, 4),
            "exact_rate":  round(exact_rate, 4),
        }

    return summary


# ---------------------------------------------------------------------------
# Print table
# ---------------------------------------------------------------------------

def print_summary_table(summary: dict) -> None:
    W = 100
    print()
    print("=" * W)
    print("  EDI Hub+ Stage 6 — Tagging Evaluation Results")
    print("=" * W)
    print(f"  {'Dimension':<30} {'TP':>5} {'TN':>5} {'FP':>5} {'FN':>5}  "
          f"{'Precision':>10} {'Recall':>8} {'F1':>8} {'Exact%':>8}")
    print("  " + "-" * (W - 2))

    overall_tp = overall_tn = overall_fp = overall_fn = 0
    overall_exact = overall_total = 0

    for dim, s in summary.items():
        label = dim.replace("_", " ").title()[:28]
        exact_pct = f"{s['exact_rate']*100:.1f}%"
        print(f"  {label:<30} {s['tp']:>5} {s['tn']:>5} {s['fp']:>5} {s['fn']:>5}  "
              f"{s['precision']:>10.4f} {s['recall']:>8.4f} {s['f1']:>8.4f} {exact_pct:>8}")
        overall_tp    += s["tp"]
        overall_tn    += s["tn"]
        overall_fp    += s["fp"]
        overall_fn    += s["fn"]
        overall_exact += s["exact_matches"]
        overall_total += s["total"]

    print("  " + "-" * (W - 2))
    prec = overall_tp / (overall_tp + overall_fp) if (overall_tp + overall_fp) > 0 else 0
    rec  = overall_tp / (overall_tp + overall_fn) if (overall_tp + overall_fn) > 0 else 0
    f1   = 2 * prec * rec / (prec + rec) if (prec + rec) > 0 else 0
    exact_pct = f"{overall_exact/overall_total*100:.1f}%" if overall_total > 0 else "—"
    print(f"  {'OVERALL':<30} {overall_tp:>5} {overall_tn:>5} {overall_fp:>5} {overall_fn:>5}  "
          f"{prec:>10.4f} {rec:>8.4f} {f1:>8.4f} {exact_pct:>8}")
    print("=" * W)
    print()


def print_detail_grid(
    eval_results:  dict,
    predictions:   dict,
    ground_truth:  dict,
    dims_filter:   list = None,
) -> None:
    """
    For each evaluated resource, prints a grid showing:
      - Each dimension as a row
      - Predicted tags vs Ground truth tags side by side
      - TP / FP / FN counts and F1 per dimension
      - Individual tags colour-coded as TP (✓) / FP (✗) / FN (missing)
    """
    dims = dims_filter or list(TAXONOMY.keys())
    DIM_W  = 28
    COL_W  = 38
    W      = DIM_W + COL_W * 2 + 30

    for rid in eval_results["matched_ids"]:
        pred = predictions.get(rid, {})
        gt   = ground_truth.get(rid, {})
        dim_scores = eval_results["per_resource"].get(rid, {})

        # Resource header
        print()
        print("┌" + "─" * (W - 2) + "┐")
        title = f"Resource {rid}"
        print(f"│  {title}".ljust(W - 1) + "│")
        print("├" + "─" * (W - 2) + "┤")
        print(f"│  {'Dimension':<{DIM_W}} │ {'Predicted':<{COL_W}} │ {'Ground Truth':<{COL_W}} │ TP  FP  FN   F1  │")
        print("├" + "─" * (W - 2) + "┤")

        for dim in dims:
            if dim not in dim_scores:
                continue

            s          = dim_scores[dim]
            pred_tags  = sorted(pred.get(dim, set()))
            gt_tags    = sorted(gt.get(dim, set()))
            gt_lower   = {t.lower() for t in gt_tags}
            pred_lower = {t.lower() for t in pred_tags}

            # Annotate predicted tags: ✓ if TP, ✗ if FP
            pred_annotated = []
            for t in pred_tags:
                marker = "✓" if t.lower() in gt_lower else "✗"
                pred_annotated.append(f"{marker}{t}")

            # Annotate ground truth tags: ✓ if matched, (missed) if FN
            gt_annotated = []
            for t in gt_tags:
                marker = "✓" if t.lower() in pred_lower else "!"
                gt_annotated.append(f"{marker}{t}")

            # Format into wrapped lines
            pred_str = ", ".join(pred_annotated) if pred_annotated else "—"
            gt_str   = ", ".join(gt_annotated)   if gt_annotated   else "—"

            # Wrap long strings
            def wrap(s, width):
                lines = []
                while len(s) > width:
                    cut = s[:width].rfind(",")
                    if cut == -1:
                        cut = width
                    else:
                        cut += 1
                    lines.append(s[:cut].strip())
                    s = s[cut:].strip()
                lines.append(s)
                return lines

            pred_lines = wrap(pred_str, COL_W)
            gt_lines   = wrap(gt_str,   COL_W)
            n_lines    = max(len(pred_lines), len(gt_lines))

            dim_label = dim.replace("_", " ").title()[:DIM_W]
            f1_str    = f"{s['f1']:.2f}"

            for i in range(n_lines):
                pl = pred_lines[i] if i < len(pred_lines) else ""
                gl = gt_lines[i]   if i < len(gt_lines)   else ""
                if i == 0:
                    counts = f"{s['tp']:>3} {s['fp']:>3} {s['fn']:>3}  {f1_str:>5}"
                    print(f"│  {dim_label:<{DIM_W}} │ {pl:<{COL_W}} │ {gl:<{COL_W}} │ {counts} │")
                else:
                    print(f"│  {'':<{DIM_W}} │ {pl:<{COL_W}} │ {gl:<{COL_W}} │ {'':>17} │")

        print("└" + "─" * (W - 2) + "┘")

    # Legend
    print()
    print("  Legend:  ✓ = correct (TP)   ✗ = false positive (FP)   ! = missed / false negative (FN)")
    print()


def print_per_resource_all_dims(eval_results: dict, dims_filter: list = None) -> None:
    """
    Print a per-resource breakdown table for every dimension —
    showing TP, FP, FN, Precision, Recall, F1 and Exact match per resource per dimension.
    """
    dims = dims_filter or list(TAXONOMY.keys())
    W    = 100

    print()
    print("=" * W)
    print("  Per-Resource Breakdown — All Dimensions")
    print("=" * W)

    for dim in dims:
        label = dim.replace("_", " ").title()
        print(f"\n  ── {label} ──")
        print(f"  {'Resource':<14} {'TP':>5} {'TN':>5} {'FP':>5} {'FN':>5} "
              f"{'Precision':>10} {'Recall':>8} {'F1':>8} {'Exact':>7}")
        print("  " + "-" * 70)

        for rid in sorted(eval_results["matched_ids"], key=lambda x: int(x) if x.isdigit() else x):
            dim_scores = eval_results["per_resource"].get(rid, {})
            if dim not in dim_scores:
                continue
            s     = dim_scores[dim]
            exact = "✓" if s["exact_match"] else "✗"
            print(f"  {rid:<14} {s['tp']:>5} {s['tn']:>5} {s['fp']:>5} {s['fn']:>5} "
                  f"{s['precision']:>10.4f} {s['recall']:>8.4f} {s['f1']:>8.4f} {exact:>7}")

    print()
    print("=" * W)
    print()


def print_tag_frequency_table(
    eval_results: dict,
    predictions:  dict,
    ground_truth: dict,
    dims_filter:  list = None,
) -> None:
    """
    For each dimension, show every tag in the taxonomy with counts of:
      TP  — correctly predicted across all resources
      FP  — predicted but wrong across all resources
      FN  — missed across all resources
      TN  — correctly not predicted across all resources
    Sorted by FP+FN (most problematic tags first).
    """
    from collections import defaultdict
    dims = dims_filter or list(TAXONOMY.keys())
    W    = 100

    print()
    print("=" * W)
    print("  Tag-Level Frequency Analysis")
    print("  (Tags ranked by FP+FN — most problematic first)")
    print("=" * W)

    for dim in dims:
        all_tags = [t.lower() for t in TAXONOMY.get(dim, [])]
        if not all_tags:
            continue

        counts = defaultdict(lambda: {"tp": 0, "fp": 0, "fn": 0, "tn": 0})

        for rid in eval_results["matched_ids"]:
            dim_scores = eval_results["per_resource"].get(rid, {})
            if dim not in dim_scores:
                continue
            per_tag = dim_scores[dim].get("per_tag", {})
            for tag, c in per_tag.items():
                counts[tag]["tp"] += c["tp"]
                counts[tag]["fp"] += c["fp"]
                counts[tag]["fn"] += c["fn"]
                counts[tag]["tn"] += c["tn"]

        # Sort by FP+FN descending
        sorted_tags = sorted(counts.items(), key=lambda x: x[1]["fp"] + x[1]["fn"], reverse=True)

        label = dim.replace("_", " ").title()
        print(f"\n  ── {label} ──")
        print(f"  {'Tag':<40} {'TP':>5} {'TN':>5} {'FP':>5} {'FN':>5}  {'Status'}")
        print("  " + "-" * 75)

        for tag, c in sorted_tags:
            if c["fp"] == 0 and c["fn"] == 0 and c["tp"] == 0:
                continue  # skip tags never seen in either predictions or ground truth
            status = []
            if c["fp"] > 0: status.append(f"{c['fp']} FP")
            if c["fn"] > 0: status.append(f"{c['fn']} FN")
            if c["tp"] > 0: status.append(f"{c['tp']} TP")
            status_str = ", ".join(status) if status else "TN only"
            print(f"  {tag.title():<40} {c['tp']:>5} {c['tn']:>5} {c['fp']:>5} {c['fn']:>5}  {status_str}")

    print()
    print("=" * W)
    print()


def print_per_resource_table(eval_results: dict, dim: str) -> None:
    """Print per-resource breakdown for a single dimension."""
    print(f"\n  Per-resource breakdown: {dim.replace('_', ' ').title()}")
    print(f"  {'Resource ID':<14} {'TP':>5} {'FP':>5} {'FN':>5} {'P':>8} {'R':>8} {'F1':>8} {'Exact'}")
    print("  " + "-" * 65)
    for rid, dims in eval_results["per_resource"].items():
        if dim not in dims:
            continue
        s     = dims[dim]
        exact = "✓" if s["exact_match"] else "✗"
        print(f"  {rid:<14} {s['tp']:>5} {s['fp']:>5} {s['fn']:>5} "
              f"{s['precision']:>8.4f} {s['recall']:>8.4f} {s['f1']:>8.4f} {exact:>6}")
    print()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _cli():
    parser = argparse.ArgumentParser(
        description="Evaluate Stage 6 tagging against ground truth Excel annotations."
    )
    parser.add_argument(
        "--tagged", nargs="+", metavar="FILE",
        help="One or more tagged_<id>.json files from stage 6",
    )
    parser.add_argument(
        "--tagged-dir", metavar="DIR",
        help="Directory to scan for tagged_*.json files",
    )
    parser.add_argument(
        "--excel", default="Data-used-in-the-prototype-12-June-2026.xlsm",
        help="Path to the ground truth Excel file",
    )
    parser.add_argument(
        "--dims", nargs="+", metavar="DIM",
        choices=list(TAXONOMY.keys()),
        help="Only evaluate these dimensions (default: all 7)",
    )
    parser.add_argument(
        "--per-resource", metavar="DIM",
        help="Also print per-resource breakdown for a specific dimension",
    )
    parser.add_argument(
        "--output-json", metavar="FILE",
        help="Save full results to a JSON file",
    )
    parser.add_argument(
        "--output-txt", metavar="FILE",
        help="Save the full console output (summary + detail grid + frequency tables) to a text file",
    )
    parser.add_argument(
        "--detail", action="store_true",
        help="Print a per-resource grid showing predicted vs ground truth tags for each dimension",
    )
    args = parser.parse_args()

    # Collect tagged files
    tagged_files = list(args.tagged or [])
    if args.tagged_dir:
        tagged_files += [str(p) for p in Path(args.tagged_dir).glob("tagged_*.json")]
    if not tagged_files:
        print("Error: no tagged files provided. Use --tagged or --tagged-dir.")
        sys.exit(1)

    print(f"\nLoading ground truth from: {args.excel}")
    ground_truth = load_ground_truth(args.excel)
    print(f"  → {len(ground_truth)} resources with ground truth annotations")

    print(f"\nLoading predictions from {len(tagged_files)} file(s)...")
    predictions = load_predictions(tagged_files)
    print(f"  → {len(predictions)} resources with predictions")

    print("\nRunning evaluation...")
    eval_results = run_evaluation(predictions, ground_truth, dims_filter=args.dims)
    print(f"  → Evaluated {len(eval_results['matched_ids'])} resource(s): {eval_results['matched_ids']}")

    summary = summarise(eval_results, dims_filter=args.dims)

    # ── Capture all output for optional text file ─────────────────────────────
    import io, contextlib

    buffer = io.StringIO()
    with contextlib.redirect_stdout(buffer):
        print(f"\nEDI Hub+ Evaluation — {len(eval_results['matched_ids'])} resource(s): {eval_results['matched_ids']}")
        print_summary_table(summary)
        if args.detail:
            print_detail_grid(eval_results, predictions, ground_truth, dims_filter=args.dims)
        print_per_resource_all_dims(eval_results, dims_filter=args.dims)
        print_tag_frequency_table(eval_results, predictions, ground_truth, dims_filter=args.dims)
        if args.per_resource:
            print_per_resource_table(eval_results, args.per_resource)

    captured = buffer.getvalue()
    print(captured)

    if args.output_txt:
        Path(args.output_txt).write_text(captured, encoding="utf-8")
        print(f"Console output saved → {args.output_txt}")

    if args.output_json:
        out = {
            "summary":      summary,
            "per_resource": eval_results["per_resource"],
            "matched_ids":  eval_results["matched_ids"],
        }
        Path(args.output_json).write_text(
            json.dumps(out, indent=2, ensure_ascii=False), encoding="utf-8"
        )
        print(f"Full results saved → {args.output_json}")


if __name__ == "__main__":
    _cli()