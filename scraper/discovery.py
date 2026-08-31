"""
Discovery — Stage 1
Responsible for:
  - Loading resource URLs from the existing prototype Excel file
  - Returning a list of Resource objects (title, author, url)

Nothing else. No fetching, no cleaning, no tagging.
"""

import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import openpyxl
from models import Resource
from utils.logger import get_logger

logger = get_logger(__name__)

SHEET_NAME    = "Resources-to-show"
COL_ID        = "Resource ID"
COL_TITLE     = "Title"
COL_AUTHOR    = "Author"
COL_URL       = "URL"


def _load_sheet(excel_path: str):
    """Open the workbook and return the resources sheet."""
    try:
        wb = openpyxl.load_workbook(
            excel_path,
            read_only=True,
            data_only=True,
        )
        if SHEET_NAME not in wb.sheetnames:
            logger.error(f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
            return None
        return wb[SHEET_NAME]
    except FileNotFoundError:
        logger.error(f"Excel file not found: {excel_path}")
        return None
    except Exception as e:
        logger.error(f"Could not open Excel file: {e}")
        return None


def discover_resources(excel_path: str) -> list[Resource]:
    """
    Entry point for Stage 1.
    Reads the Excel prototype data and returns all Resource objects
    that have a valid URL.
    """
    logger.info(f"Loading resources from Excel: {excel_path}")

    ws = _load_sheet(excel_path)
    if ws is None:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        logger.error("Sheet is empty")
        return []

    # First row is headers
    headers = [str(h).strip() if h else "" for h in rows[0]]
    logger.debug(f"Columns found: {headers}")

    # Validate required columns exist
    for required in [COL_TITLE, COL_URL]:
        if required not in headers:
            logger.error(f"Required column '{required}' not found in sheet")
            return []

    resources = []
    skipped   = 0

    for row in rows[1:]:
        if not row or not row[0]:
            continue  # skip empty rows

        data = dict(zip(headers, row))

        url = str(data.get(COL_URL, "")).strip()
        if not url or not url.startswith("http"):
            skipped += 1
            logger.debug(f"Skipping row — no valid URL: {data.get(COL_TITLE, '?')}")
            continue

        resource = Resource(
            title       = str(data.get(COL_TITLE,  "Unknown")).strip(),
            url         = url,
            author      = str(data.get(COL_AUTHOR, "Unknown")).strip(),
            resource_id = str(data.get(COL_ID,     "")).strip(),
        )
        resources.append(resource)
        logger.debug(f"  [{resource.resource_id}] {resource.title[:60]}")

    logger.info(f"Stage 1 complete — {len(resources)} resources loaded, {skipped} skipped (no URL)")
    return resources